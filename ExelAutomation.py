import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def process_workbook(file_name):
    wb = xl.load_workbook('file_name')
    sheet = wb['Sheet1']

    for i in range(2, sheet.max_row + 1):
        x = sheet.cell(i, 3)
        z = 0.9 * x.value
        n_cell = sheet.cell(i, 4)
        n_cell.value = z

    # for row in range(2, sheet.max_row + 1):
    #     y = sheet.cell(row,4)
    v = Reference(sheet,
                  min_row=2,
                  max_row=sheet.max_row,
                  min_col=4,
                  max_col=4)
    chart = BarChart()
    chart.add_data(v)
    sheet.add_chart(chart, 'e2')
    wb.save('file_name')
